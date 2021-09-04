from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from rcs_back.containers_app.models import Container


def format_td(td: timedelta) -> str:
    td_s = str(td)
    if "day" in td_s:
        td_s = td_s.split(":")[0]
        days = td_s[:td_s.find(" ")]
        hours = td_s[td_s.find(",") + 2:]
        return f"{days} дн {hours} ч"
    else:
        return td_s.split(":")[0] + " ч"


def write_headers(ws: Worksheet) -> None:
    """Записывает заголовки в статистику о контейнере"""
    ws["A2"] = "ID"
    ws["B2"] = "Вид контейнера"
    ws["C2"] = "Масса, кг"
    ws["D2"] = "Адрес здания"
    ws["E2"] = "Номер корпуса"
    ws["F2"] = "Этаж"
    ws["G2"] = "Аудитория"
    ws["H2"] = "Описание"
    ws["I2"] = "Состояние"
    ws["J2"] = "Текущее"
    ws["K2"] = "Среднее"
    ws["L2"] = "Текущее"
    ws["M2"] = "Среднее"
    ws["N2"] = "Номер телефона"
    ws["O2"] = "Почта"
    ws["P2"] = "Суммарная масса, кг"
    ws["J1"] = "Время заполнения"
    ws["L1"] = "Ожидание сбора после заполнения"
    ws["N1"] = "Контактное лицо"

    ws.row_dimensions[1].font = Font(bold=True)
    ws.row_dimensions[2].font = Font(bold=True)


def write_container(c: Container, ws: Worksheet, i: int) -> None:
    """Записывает статистику одного контейнера"""
    ws[f"A{i}"] = c.pk
    ws[f"B{i}"] = c.get_kind_display()
    ws[f"C{i}"] = c.mass()
    ws[f"D{i}"] = c.building.address
    if c.building_part:
        ws[f"E{i}"] = c.building_part.num
    else:
        ws[f"E{i}"] = "-"
    ws[f"F{i}"] = c.floor
    ws[f"G{i}"] = c.room
    ws[f"H{i}"] = c.description
    ws[f"I{i}"] = c.get_status_display()
    if c.cur_fill_time():
        ws[f"J{i}"] = format_td(c.cur_fill_time())
    else:
        ws[f"J{i}"] = "Уже заполнен"
    if c.avg_fill_time:
        ws[f"K{i}"] = format_td(c.avg_fill_time)
    else:
        ws[f"K{i}"] = "Недостаточно данных"
    if c.cur_takeout_wait_time():
        ws[f"L{i}"] = format_td(c.cur_takeout_wait_time())
    else:
        ws[f"L{i}"] = "Ещё не заполнен"
    if c.avg_takeout_wait_time:
        ws[f"M{i}"] = format_td(c.avg_takeout_wait_time)
    else:
        ws[f"M{i}"] = "Недостаточно данных"
    ws[f"N{i}"] = c.phone
    ws[f"O{i}"] = c.email
    ws[f"P{i}"] = c.collected_mass()


def set_width(ws: Worksheet) -> None:
    """Ширина столбца как наибольшая из клеток"""
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0),
                     len(str(cell.value)))
                ) + 0.5
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def get_container_stats_xl() -> Workbook:
    """Создаёт excel-файл с актуальной статистикой по контейнерам"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Контейнеры"
    write_headers(ws)
    containers = Container.objects.order_by("pk")

    for i in range(3, len(containers) + 3):
        container = containers[i-3]
        write_container(container, ws, i)

    set_width(ws)

    return wb
