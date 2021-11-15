from datetime import timedelta

from django.db.models import Q
from django.db.models.query import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from rcs_back.containers_app.models import Building, Container
from rcs_back.takeouts_app.models import ContainersTakeoutRequest, TankTakeoutRequest


def queryset_to_ids(qs: QuerySet) -> str:
    """Форматирование QuerySet"""
    string = ""
    for item in qs:
        string += str(item.pk)
        string += ", "
    return string[:len(string)-2]


def format_td(timed: timedelta) -> str:
    """Форматирование timedelta"""
    td_s = str(timed)
    if "day" in td_s:
        td_s = td_s.split(":", maxsplit=1)[0]
        days = td_s[:td_s.find(" ")]
        hours = td_s[td_s.find(",") + 2:]
        return f"{days} дн {hours} ч"
    else:
        return td_s.split(":", maxsplit=1)[0] + " ч"


def set_width(worksheet: Worksheet) -> None:
    """Ширина столбца как наибольшая из клеток"""
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0),
                     len(str(cell.value)))
                )
                if dims[cell.column_letter] > 20:
                    dims[cell.column_letter] = 30
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value + 2


def write_container_headers(worksheet: Worksheet) -> None:
    """Записывает заголовки в статистику о контейнере"""
    worksheet["A2"] = "ID"
    worksheet["B2"] = "Вид контейнера"
    worksheet["C2"] = "Масса, кг"
    worksheet["D2"] = "Адрес здания"
    worksheet["E2"] = "Номер корпуса"
    worksheet["F2"] = "Этаж"
    worksheet["G2"] = "Аудитория"
    worksheet["H2"] = "Описание"
    worksheet["I2"] = "Состояние"
    worksheet["J2"] = "Текущее"
    worksheet["K2"] = "Среднее"
    worksheet["L2"] = "Текущее"
    worksheet["M2"] = "Среднее"
    worksheet["N2"] = "Номер телефона"
    worksheet["O2"] = "Почта"
    worksheet["P2"] = "Суммарная масса, кг"
    worksheet["J1"] = "Время заполнения"
    worksheet["L1"] = "Ожидание сбора после заполнения"
    worksheet["N1"] = "Контактное лицо"

    worksheet.row_dimensions[1].font = Font(bold=True)
    worksheet.row_dimensions[2].font = Font(bold=True)


def write_container(container: Container, worksheet: Worksheet, i: int) -> None:
    """Записывает статистику одного контейнера"""
    worksheet[f"A{i}"] = container.pk
    worksheet[f"B{i}"] = container.get_kind_display()
    worksheet[f"C{i}"] = container.mass()
    worksheet[f"D{i}"] = container.building.address
    if container.building_part:
        worksheet[f"E{i}"] = container.building_part.num
    else:
        worksheet[f"E{i}"] = "-"
    worksheet[f"F{i}"] = container.floor
    worksheet[f"G{i}"] = container.room
    worksheet[f"H{i}"] = container.description
    worksheet[f"I{i}"] = container.get_status_display()
    if container.cur_fill_time():
        worksheet[f"J{i}"] = format_td(container.cur_fill_time())
    elif container.is_active():
        worksheet[f"J{i}"] = "Уже заполнен"
    else:
        worksheet[f"J{i}"] = "Ещё не подключён"
    if container.avg_fill_time:
        worksheet[f"K{i}"] = format_td(container.avg_fill_time)
    else:
        worksheet[f"K{i}"] = "Недостаточно данных"
    if container.cur_takeout_wait_time():
        worksheet[f"L{i}"] = format_td(container.cur_takeout_wait_time())
    else:
        worksheet[f"L{i}"] = "Ещё не заполнен"
    if container.avg_takeout_wait_time:
        worksheet[f"M{i}"] = format_td(container.avg_takeout_wait_time)
    else:
        worksheet[f"M{i}"] = "Недостаточно данных"
    worksheet[f"N{i}"] = container.phone
    worksheet[f"O{i}"] = container.email
    worksheet[f"P{i}"] = container.collected_mass()


def get_container_stats_ws(worksheet: Worksheet) -> None:
    """Создаёт страницу из excel с актуальной статистикой по контейнерам"""
    worksheet.title = "Контейнеры"
    write_container_headers(worksheet)
    containers = Container.objects.filter(
        ~Q(status=Container.RESERVED)
    ).order_by("pk")

    for i in range(3, len(containers) + 3):
        container = containers[i-3]
        write_container(container, worksheet, i)

    set_width(worksheet)


def get_container_stats_xl() -> Workbook:
    """Создаёт excel-WorkBook с актуальной статистикой по контейнерам"""
    workbook = Workbook()
    worksheet = workbook.active
    get_container_stats_ws(worksheet)
    return workbook


def write_container_takeout_headers(worksheet: Worksheet) -> None:
    """Записывает заголовки в статистику о сборе"""
    worksheet["A1"] = "Дата сбора"
    worksheet["A2"] = "создание"
    worksheet["B2"] = "подтверждение"
    worksheet["C1"] = "Здание"
    worksheet["D1"] = "Корпус"
    worksheet["E1"] = "Список контейнеров на сбор"
    worksheet["F1"] = "Неподтверждённые контейнеры"
    worksheet["G1"] = "Соответствие"
    worksheet["H1"] = "Суммарная масса сбора"
    worksheet["I1"] = "Данные подсобного рабочего"

    worksheet.row_dimensions[1].font = Font(bold=True)
    worksheet.row_dimensions[2].font = Font(bold=True)


def write_container_takeout(request: ContainersTakeoutRequest,
                            worksheet: Worksheet, i: int) -> None:
    """Записывает статистику одного сбора"""
    worksheet[f"A{i}"] = request.created_at.strftime("%d.%m.%Y")
    if request.confirmed_at:
        worksheet[f"B{i}"] = request.confirmed_at.strftime("%d.%m.%Y")
    else:
        worksheet[f"B{i}"] = "-"
    worksheet[f"C{i}"] = request.building.address
    if request.building_part:
        worksheet[f"D{i}"] = request.building_part.num
    else:
        worksheet[f"D{i}"] = "-"
    worksheet[f"E{i}"] = queryset_to_ids(request.containers.all())
    worksheet[f"F{i}"] = queryset_to_ids(request.unconfirmed_containers())
    worksheet[f"G{i}"] = request.emptied_containers_match()
    worksheet[f"H{i}"] = request.mass()
    worksheet[f"I{i}"] = request.worker_info


def get_container_takeout_stats_ws(worksheet: Worksheet) -> None:
    """Создаёт страницу из excel с актуальной статистикой по сборам"""
    worksheet.title = "Сборы"
    write_container_takeout_headers(worksheet)
    container_takeouts = ContainersTakeoutRequest.objects.order_by(
        "pk")

    for i in range(3, len(container_takeouts) + 3):
        takeout = container_takeouts[i-3]
        write_container_takeout(takeout, worksheet, i)

    set_width(worksheet)


def get_container_takeout_stats_xl() -> Workbook:
    """Создаёт excel-WorkBook с актуальной статистикой по сборам"""
    workbook = Workbook()
    worksheet = workbook.active
    get_container_takeout_stats_ws(worksheet)
    return workbook


def write_tank_takeout_headers(worksheet: Worksheet) -> None:
    """Записывает заголовки в статистику о вывозе"""
    worksheet["A1"] = "Дата обращения к оператору"
    worksheet["B1"] = "Дата вывоза"
    worksheet["C1"] = "Здание"
    worksheet["D1"] = "Время заполнения накопительного бака"
    worksheet["E1"] = "Расчётная масса вывоза, кг"
    worksheet["F1"] = "Подтверждённая масса вывоза, кг"
    worksheet["G1"] = "Соответствие (расчётная/подтверждённая"
    worksheet["H1"] = "Разница (расчётная - подтверждённая), кг"

    worksheet.row_dimensions[1].font = Font(bold=True)


def write_tank_takeout(request: TankTakeoutRequest,
                       worksheet: Worksheet, i: int) -> None:
    """Записывает статистику одного вывоза"""
    worksheet[f"A{i}"] = request.created_at.strftime("%d.%m.%Y")
    if request.confirmed_at:
        worksheet[f"B{i}"] = request.confirmed_at.strftime("%d.%m.%Y")
    else:
        worksheet[f"B{i}"] = "-"
    worksheet[f"C{i}"] = request.building.address
    if request.fill_time():
        worksheet[f"D{i}"] = format_td(request.fill_time())
    else:
        worksheet[f"D{i}"] = "Недостаточно данных"
    worksheet[f"E{i}"] = request.mass()
    worksheet[f"F{i}"] = request.confirmed_mass
    worksheet[f"G{i}"] = request.confirmed_mass_match()
    worksheet[f"H{i}"] = request.mass_difference()


def get_tank_takeout_stats_ws(worksheet: Worksheet) -> None:
    """Создаёт страницу из excel с актуальной статистикой по вывозам"""
    worksheet.title = "Вывозы"
    write_tank_takeout_headers(worksheet)
    tank_takeouts = TankTakeoutRequest.objects.order_by(
        "pk")

    for i in range(2, len(tank_takeouts) + 2):
        takeout = tank_takeouts[i-2]
        write_tank_takeout(takeout, worksheet, i)

    set_width(worksheet)


def write_building_headers(worksheet: Worksheet) -> None:
    """Записывает заголовки в статистику по зданию"""
    worksheet["A1"] = "Здание"
    worksheet["B1"] = "Число контейнеров"
    worksheet["C1"] = "Суммарный объём, кг"
    worksheet["D1"] = "Средняя скорость сбора, кг/месяц"

    worksheet.row_dimensions[1].font = Font(bold=True)


def write_building(building: Building, worksheet: Worksheet, i: int) -> None:
    """Записывает статистику одного здания"""
    worksheet[f"A{i}"] = building.address
    worksheet[f"B{i}"] = building.container_count()
    worksheet[f"C{i}"] = building.confirmed_collected_mass()
    worksheet[f"D{i}"] = building.avg_fill_speed()


def get_building_stats_ws(worksheet: Worksheet) -> None:
    """Создаёт страницу из excel с актуальной статистикой по зданию"""
    worksheet.title = "По зданиям"
    write_building_headers(worksheet)
    buildings = Building.objects.order_by("pk")

    for i in range(2, len(buildings) + 2):
        building = buildings[i-2]
        write_building(building, worksheet, i)

    set_width(worksheet)


def get_tank_takeout_stats_xl() -> Workbook:
    """Создаёт excel-WorkBook с актуальной статистикой по вывозам"""
    workbook = Workbook()
    worksheet = workbook.active
    get_tank_takeout_stats_ws(worksheet)
    ws1 = workbook.create_sheet()
    get_building_stats_ws(ws1)
    return workbook


def get_all_stats_xl() -> Workbook:
    """Создаёт excel-WorkBook со всей актуальной статистикой"""
    workbook = Workbook()
    worksheet = workbook.active
    get_container_stats_ws(worksheet)
    ws1 = workbook.create_sheet()
    get_container_takeout_stats_ws(ws1)
    ws2 = workbook.create_sheet()
    get_tank_takeout_stats_ws(ws2)
    ws3 = workbook.create_sheet()
    get_building_stats_ws(ws3)
    return workbook


def write_short_container_headers(worksheet: Worksheet) -> None:
    """Записывает заголовки в короткую статистику о контейнерах"""
    worksheet["A1"] = "ID"
    worksheet["B1"] = "Вид контейнера"
    worksheet["C1"] = "Номер корпуса"
    worksheet["D1"] = "Этаж"
    worksheet["E1"] = "Аудитория"
    worksheet["F1"] = "Описание"
    worksheet["G1"] = "Номер телефона"
    worksheet["H1"] = "Почта"

    worksheet.row_dimensions[1].font = Font(bold=True)


def write_short_container_info(container: Container, worksheet: Worksheet, i: int) -> None:
    """Записывает краткую информацию об одном контейнере"""
    worksheet[f"A{i}"] = container.pk
    worksheet[f"B{i}"] = container.get_kind_display()
    if container.building_part:
        worksheet[f"C{i}"] = container.building_part.num
    else:
        worksheet[f"C{i}"] = "-"
    worksheet[f"D{i}"] = container.floor
    worksheet[f"E{i}"] = container.room
    worksheet[f"F{i}"] = container.description
    worksheet[f"G{i}"] = container.phone
    worksheet[f"H{i}"] = container.email


def get_short_container_info_ws(worksheet: Worksheet,
                                containers: QuerySet[Container]) -> None:
    """Создаёт страницу из excel с краткой информацией
    по выбранным контейнерам"""
    worksheet.title = "Контейнеры"
    write_short_container_headers(worksheet)

    for i in range(2, len(containers) + 2):
        container = containers[i-2]
        write_short_container_info(container, worksheet, i)

    set_width(worksheet)


def get_short_container_info_xl(containers: QuerySet[Container]) -> Workbook:
    """Создаёт excel-WorkBook с актуальной статистикой по контейнерам"""
    workbook = Workbook()
    worksheet = workbook.active
    get_short_container_info_ws(worksheet, containers)
    return workbook
